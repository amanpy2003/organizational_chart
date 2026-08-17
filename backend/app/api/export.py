import io

from fastapi import APIRouter, HTTPException, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.schemas.employee import OrgNodeSchema
from app.schemas.export import ExportPdfRequest
from app.services.pdf_service import generate_org_chart_pdf

router = APIRouter()


@router.post("/export/pdf")
def export_pdf(request: ExportPdfRequest) -> Response:
    if not request.trees:
        raise HTTPException(status_code=400, detail="No organization data provided to export.")

    try:
        pdf_bytes = generate_org_chart_pdf(
            trees=request.trees,
            fields=request.fields,
            direction=request.direction,
            page_size=request.page_size,
            orientation=request.orientation,
            title=request.title,
            scope_label=request.scope_label,
            fit_to_one_page=request.fit_to_one_page,
        )
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Failed to generate PDF: {exc}") from exc

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="organization_chart.pdf"'},
    )


@router.post("/export/excel")
def export_excel(request: ExportPdfRequest) -> Response:
    """Re-emit the (possibly filtered) current dataset as a flat .xlsx sheet."""
    if not request.trees:
        raise HTTPException(status_code=400, detail="No organization data provided to export.")

    rows: list[OrgNodeSchema] = []

    def walk(node: OrgNodeSchema) -> None:
        rows.append(node)
        for child in node.children:
            walk(child)

    for tree in request.trees:
        walk(tree)

    wb = Workbook()
    ws = wb.active
    ws.title = "Organization Data"
    headers = [
        "Employee ID",
        "Employee Name",
        "Designation",
        "Department",
        "Location",
        "Email",
        "Level",
        "Employment Type",
        "Status",
    ]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, node in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=node.employee_id)
        ws.cell(row=row_idx, column=2, value=node.name)
        ws.cell(row=row_idx, column=3, value=node.designation)
        ws.cell(row=row_idx, column=4, value=node.department)
        ws.cell(row=row_idx, column=5, value=node.location)
        ws.cell(row=row_idx, column=6, value=node.email)
        ws.cell(row=row_idx, column=7, value=node.level)
        ws.cell(row=row_idx, column=8, value=node.employment_type)
        ws.cell(row=row_idx, column=9, value=node.status)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="organization_export.xlsx"'},
    )
