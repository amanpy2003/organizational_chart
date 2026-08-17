"""Builds the downloadable Organization Chart Excel template in-memory.

Kept as the single source of truth for the expected column schema so the
template can never drift out of sync with excel_service's HEADER_ALIASES.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

COLUMNS = [
    ("Employee ID", 16),
    ("Employee Name", 26),
    ("Designation", 24),
    ("Department", 20),
    ("Reports To Employee ID", 22),
    ("Location", 18),
    ("Email", 28),
    ("Level", 10),
    ("Employment Type", 18),
    ("Status", 12),
]

SAMPLE_ROWS = [
    ["E001", "Amit Sharma", "Managing Director", "Corporate", "", "Mumbai", "amit.sharma@company.com", "L1", "Full-Time", "Active"],
    ["E002", "Rahul Verma", "VP Projects", "Projects", "E001", "Mumbai", "rahul.verma@company.com", "L2", "Full-Time", "Active"],
    ["E003", "Neha Singh", "VP Finance", "Finance", "E001", "Delhi", "neha.singh@company.com", "L2", "Full-Time", "Active"],
    ["E004", "Raj Kumar", "Project Director", "Projects", "E002", "Mumbai", "raj.kumar@company.com", "L3", "Full-Time", "Active"],
    ["E005", "Ankit Jain", "Project Manager", "Projects", "E002", "Pune", "ankit.jain@company.com", "L3", "Full-Time", "Active"],
    ["E006", "Priya Mehta", "Finance Manager", "Finance", "E003", "Delhi", "priya.mehta@company.com", "L3", "Full-Time", "Active"],
]

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
BODY_FONT = Font(size=10.5, name="Calibri")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def build_template_workbook() -> bytes:
    wb = Workbook()

    _build_data_sheet(wb)
    _build_instructions_sheet(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_data_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Organization Data"

    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, row_values in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    status_validation = DataValidation(
        type="list", formula1='"Active,Inactive,On Leave"', allow_blank=True
    )
    status_validation.error = "Please choose a value from the list."
    status_validation.errorTitle = "Invalid Status"
    ws.add_data_validation(status_validation)
    status_validation.add("J2:J1001")

    employment_validation = DataValidation(
        type="list",
        formula1='"Full-Time,Part-Time,Contract,Intern"',
        allow_blank=True,
    )
    employment_validation.error = "Please choose a value from the list."
    employment_validation.errorTitle = "Invalid Employment Type"
    ws.add_data_validation(employment_validation)
    employment_validation.add("I2:I1001")

    ws.sheet_view.showGridLines = True


def _build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90

    title_cell = ws.cell(row=1, column=1, value="How to fill this template")
    title_cell.font = Font(bold=True, size=14, name="Calibri", color="1E3A5F")
    ws.merge_cells("A1:B1")

    rows = [
        ("Employee ID", "Required. A unique identifier for every employee (e.g. E001). Never reuse an ID."),
        ("Employee Name", "Required. Full name of the employee."),
        ("Designation", "Job title / designation. Descriptive only — does not affect hierarchy."),
        ("Department", "Department or function. Descriptive only — does not affect hierarchy."),
        (
            "Reports To Employee ID",
            "The Employee ID of this person's direct manager. Leave BLANK for the top-most "
            "person(s) in the organization. This is the ONLY field used to build the hierarchy.",
        ),
        ("Location", "Optional. Office location or city."),
        ("Email", "Optional. Work email address."),
        ("Level", "Optional. A descriptive level/band label — does not affect hierarchy."),
        ("Employment Type", "Optional. Full-Time, Part-Time, Contract, or Intern."),
        ("Status", "Optional. Active, Inactive, or On Leave."),
        ("", ""),
        (
            "Important",
            "Do not use Employee Name to establish reporting relationships — always use "
            "Employee ID, since names can repeat. The application will reject the file if "
            "Employee IDs are missing/duplicated, if a Reports-To ID does not exist, or if a "
            "circular reporting relationship is detected.",
        ),
    ]

    for row_idx, (label, description) in enumerate(rows, start=3):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True, size=10.5, name="Calibri")
        label_cell.alignment = Alignment(vertical="top", wrap_text=True)

        desc_cell = ws.cell(row=row_idx, column=2, value=description)
        desc_cell.font = Font(size=10.5, name="Calibri")
        desc_cell.alignment = Alignment(vertical="top", wrap_text=True)
