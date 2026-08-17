"""Generates the sample/test Excel fixtures used to exercise the application.

Run with:  python scripts/generate_sample_data.py
Writes .xlsx files into backend/sample_data/.
"""
from __future__ import annotations

import random
from collections import deque
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUT_DIR = Path(__file__).resolve().parent.parent / "sample_data"

HEADERS = [
    "Employee ID",
    "Employee Name",
    "Designation",
    "Department",
    "Reports To Employee ID",
    "Location",
    "Email",
    "Level",
    "Employment Type",
    "Status",
]

FIRST_NAMES = [
    "Amit", "Rahul", "Neha", "Priya", "Raj", "Ankit", "Suresh", "Kavita", "Vikram", "Anjali",
    "Sanjay", "Pooja", "Arjun", "Divya", "Manoj", "Ritu", "Karan", "Sneha", "Deepak", "Meera",
    "Rohit", "Nisha", "Vivek", "Shalini", "Aditya", "Swati", "Gaurav", "Preeti", "Nikhil", "Anita",
    "Venkataramanaiah", "Chandrasekaran", "Rajalakshmi", "Balasubramaniam",
]
LAST_NAMES = [
    "Sharma", "Verma", "Singh", "Mehta", "Kumar", "Jain", "Gupta", "Rao", "Nair", "Iyer",
    "Reddy", "Patel", "Choudhury", "Bose", "Malhotra", "Kapoor", "Chatterjee", "Desai", "Pillai",
    "Subramaniam", "Krishnan",
]
LOCATIONS = ["Mumbai", "Delhi", "Bengaluru", "Pune", "Chennai", "Hyderabad", "Kolkata", "Ahmedabad", "Gurugram"]
DEPARTMENTS = [
    "Corporate", "Projects", "Finance", "Human Resources", "Sales", "Information Technology",
    "Legal", "Marketing", "Operations", "Procurement",
]
LEVEL_TITLES = [
    "Managing Director", "Vice President", "Director", "Senior Manager", "Manager",
    "Team Lead", "Senior Associate", "Associate", "Analyst", "Executive",
]
EMPLOYMENT_TYPES = ["Full-Time", "Full-Time", "Full-Time", "Part-Time", "Contract", "Intern"]
STATUSES = ["Active", "Active", "Active", "Active", "On Leave", "Inactive"]


def _style_header(ws) -> None:
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        ws.column_dimensions[cell.column_letter].width = max(16, len(header) + 2)


def write_workbook(filename: str, rows: list[list]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Organization Data"
    _style_header(ws)
    for row in rows:
        ws.append(row)
    path = OUT_DIR / filename
    wb.save(path)
    print(f"wrote {path} ({len(rows)} employees)")


def _row(emp_id, name, designation, department, reports_to, location, email, level, emp_type, status):
    return [emp_id, name, designation, department, reports_to, location, email, level, emp_type, status]


def _email(name: str, emp_id: str) -> str:
    slug = name.lower().replace(" ", ".")
    return f"{slug}.{emp_id.lower()}@company.example"


def generate_org(
    total: int,
    seed: int,
    branching: list[int],
    wide_manager_extra: int = 0,
    deep_chain_length: int = 0,
) -> list[list]:
    """Builds a plausible org tree via breadth-first expansion.

    branching[d] = typical number of direct reports for a manager at depth d
    (depth of the manager, i.e. children land at depth d+1). Depths beyond the
    list length taper down to 2.
    """
    rng = random.Random(seed)
    rows: list[list] = []
    next_id = 1

    def new_id() -> str:
        nonlocal next_id
        emp_id = f"E{next_id:04d}"
        next_id += 1
        return emp_id

    root_id = new_id()
    root_name = f"{rng.choice(FIRST_NAMES)} {rng.choice(LAST_NAMES)}"
    rows.append(
        _row(root_id, root_name, LEVEL_TITLES[0], "Corporate", "", rng.choice(LOCATIONS),
             _email(root_name, root_id), "L1", "Full-Time", "Active")
    )

    # queue entries: (manager_id, depth, department)
    queue: deque = deque([(root_id, 0, "Corporate")])
    wide_target_assigned = False
    leaf_ids: list[str] = [root_id]
    reserve_for_deep_chain = deep_chain_length

    while queue and len(rows) < total - reserve_for_deep_chain:
        manager_id, depth, department = queue.popleft()
        child_depth = depth + 1
        branch_count = branching[depth] if depth < len(branching) else max(2, branching[-1] // 2)

        if wide_manager_extra and not wide_target_assigned and depth == 2:
            branch_count = wide_manager_extra
            wide_target_assigned = True

        for _ in range(branch_count):
            if len(rows) >= total - reserve_for_deep_chain:
                break
            emp_id = new_id()
            name = f"{rng.choice(FIRST_NAMES)} {rng.choice(LAST_NAMES)}"
            dept = department if depth > 0 else rng.choice(DEPARTMENTS)
            title = LEVEL_TITLES[min(child_depth, len(LEVEL_TITLES) - 1)]
            emp = _row(
                emp_id, name, title, dept, manager_id, rng.choice(LOCATIONS),
                _email(name, emp_id), f"L{child_depth + 1}",
                rng.choice(EMPLOYMENT_TYPES), rng.choice(STATUSES),
            )
            rows.append(emp)
            leaf_ids.append(emp_id)
            if child_depth < len(branching) + 2:
                queue.append((emp_id, child_depth, dept))

    if deep_chain_length:
        chain_parent = leaf_ids[-1]
        chain_dept = rows[-1][3] if rows else "Corporate"
        for i in range(deep_chain_length):
            emp_id = new_id()
            name = f"{rng.choice(FIRST_NAMES)} {rng.choice(LAST_NAMES)}"
            title = f"Specialist Level {i + 1}"
            emp = _row(
                emp_id, name, title, chain_dept, chain_parent, rng.choice(LOCATIONS),
                _email(name, emp_id), f"D{i + 1}", "Full-Time", "Active",
            )
            rows.append(emp)
            chain_parent = emp_id

    return rows


def generate_valid_small() -> list[list]:
    rng = random.Random(7)
    rows: list[list] = []

    def add(emp_id, name, title, dept, manager, location):
        rows.append(_row(emp_id, name, title, dept, manager, location, _email(name, emp_id),
                          f"L{len(rows)}", "Full-Time", "Active"))

    add("E001", "Amit Sharma", "Managing Director", "Corporate", "", "Mumbai")
    add("E002", "Rahul Verma", "VP Projects", "Projects", "E001", "Mumbai")
    add("E003", "Neha Singh", "VP Finance", "Finance", "E001", "Delhi")
    add("E004", "Karan Malhotra", "VP Human Resources", "Human Resources", "E001", "Gurugram")
    add("E005", "Sneha Iyer", "VP Sales", "Sales", "E001", "Bengaluru")
    add("E006", "Venkataramanaiah Chandrasekaran", "VP Information Technology & Digital Transformation", "Information Technology", "E001", "Chennai")

    # Wide fan-out under VP Projects (12 direct reports) to exercise the
    # "many direct reports" requirement even in the small sample.
    project_manager_ids = []
    for i in range(12):
        emp_id = f"E{100 + i}"
        name = f"{rng.choice(FIRST_NAMES)} {rng.choice(LAST_NAMES)}"
        add(emp_id, name, "Project Manager", "Projects", "E002", rng.choice(LOCATIONS))
        project_manager_ids.append(emp_id)

    # A handful of individual contributors under the first three project managers.
    for j, pm_id in enumerate(project_manager_ids[:3]):
        for k in range(4):
            emp_id = f"E{200 + j * 4 + k}"
            name = f"{rng.choice(FIRST_NAMES)} {rng.choice(LAST_NAMES)}"
            add(emp_id, name, "Project Engineer", "Projects", pm_id, rng.choice(LOCATIONS))

    for dept, mgr, prefix in [
        ("Finance", "E003", 300), ("Human Resources", "E004", 320),
        ("Sales", "E005", 340), ("Information Technology", "E006", 360),
    ]:
        for i in range(4):
            emp_id = f"E{prefix + i}"
            name = f"{rng.choice(FIRST_NAMES)} {rng.choice(LAST_NAMES)}"
            add(emp_id, name, "Manager", dept, mgr, rng.choice(LOCATIONS))

    return rows


def main() -> None:
    write_workbook("valid_small_50.xlsx", generate_valid_small())

    write_workbook(
        "valid_large_500.xlsx",
        generate_org(
            total=500,
            seed=42,
            branching=[10, 4, 5, 4, 3, 2],
            wide_manager_extra=35,
            deep_chain_length=10,
        ),
    )

    # --- Bad-data fixtures, each derived from a small valid base ---
    base = generate_valid_small()[:15]

    duplicate_rows = [list(r) for r in base]
    duplicate_rows.append(list(base[3]))  # duplicate E004 row
    write_workbook("invalid_duplicate_ids.xlsx", duplicate_rows)

    missing_rows = [list(r) for r in base]
    missing_rows[5][0] = ""  # blank Employee ID, Reports-To of others still points at E006 (now orphaned on purpose is avoided)
    write_workbook("invalid_missing_ids.xlsx", missing_rows)

    invalid_manager_rows = [list(r) for r in base]
    invalid_manager_rows[7][4] = "E999"  # points at a manager that doesn't exist
    write_workbook("invalid_manager_reference.xlsx", invalid_manager_rows)

    circular_rows = [list(r) for r in base]
    # Break the E002->E001 link and create a 3-node cycle: E002 -> E003 -> E004 -> E002
    for row in circular_rows:
        if row[0] == "E002":
            row[4] = "E004"
        elif row[0] == "E003":
            row[4] = "E002"
        elif row[0] == "E004":
            row[4] = "E003"
    write_workbook("invalid_circular_reference.xlsx", circular_rows)

    multi_root_rows = [list(r) for r in base]
    for row in multi_root_rows:
        if row[0] in ("E003", "E004"):
            row[4] = ""  # give two more employees a blank Reports-To
    write_workbook("invalid_multiple_roots.xlsx", multi_root_rows)


if __name__ == "__main__":
    main()
